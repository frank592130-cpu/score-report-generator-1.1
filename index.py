from __future__ import annotations

import io
import json
import logging
import math
import re

import openpyxl
from flask import Flask, jsonify, render_template_string, request, send_file
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.exceptions import InvalidFileException


app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024

LOGGER = logging.getLogger(__name__)
X_MAX = 25.0
Y_MAX = 6.0
X_WEIGHT = 85.0
Y_WEIGHT = 15.0
THRESHOLD_KEYS = ("th_app", "th_ap", "th_a", "th_bpp")
GRADE_THRESHOLDS = (
    ("A++", "th_app"),
    ("A+", "th_ap"),
    ("A", "th_a"),
    ("B++", "th_bpp"),
)
XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# Core Excel drawing logic
FONT_NAME = "新細明體"
FILLS = {
    "A++": None,
    "A+": PatternFill("solid", fgColor="E6E6E6"),
    "A": PatternFill("solid", fgColor="BFBFBF"),
    "B++": PatternFill("solid", fgColor="808080"),
    "": PatternFill("solid", fgColor="808080"),
}
REPORT_COLUMN_WIDTHS = {
    "A": 7.25,
    "E": 7.25,
    "I": 7.25,
    "B": 4.50,
    "C": 4.50,
    "F": 4.50,
    "G": 4.50,
    "D": 5.75,
    "H": 5.75,
    "J": 8.88,
    "K": 8.88,
    "L": 8.88,
    "M": 0.4,
    "N": 7,
    "O": 7,
    "P": 7,
}


def _med():
    return Side(style="medium", color="000000")


def _thn():
    return Side(style="thin", color="000000")


def all_thin():
    s = _thn()
    return Border(left=s, right=s, top=s, bottom=s)


def outer_med(r, c, r1, c1, r2, c2):
    return Border(
        left=_med() if c == c1 else _thn(),
        right=_med() if c == c2 else _thn(),
        top=_med() if r == r1 else _thn(),
        bottom=_med() if r == r2 else _thn(),
    )


def sc(ws, row, col, value, bold=False, size=10, fill=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=FONT_NAME, bold=bold, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if border:
        cell.border = border
    if fill:
        cell.fill = fill
    return cell


def apply_report_column_widths(ws):
    for col_letter, width in REPORT_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def to_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def clamp(value, low, high):
    return min(high, max(low, value))


def total_score(x, y):
    return (x / X_MAX) * X_WEIGHT + (y / Y_MAX) * Y_WEIGHT


def grade_for_score(total, ths):
    for grade, key in GRADE_THRESHOLDS:
        if total >= ths[key]:
            return grade
    return ""


def threshold_values(source):
    return {key: to_float(source.get(key, 0)) for key in THRESHOLD_KEYS}


def normalize_student(raw):
    if not isinstance(raw, dict):
        return None

    name = str(raw.get("name", "")).strip()
    if not name:
        return None

    is_leave = bool(raw.get("is_leave", False))
    x = 0.0 if is_leave else clamp(to_float(raw.get("x", 0)), 0.0, X_MAX)
    y = 0.0 if is_leave else clamp(to_float(raw.get("y", 0)), 0.0, Y_MAX)

    return {
        "id": str(raw.get("id", "")).strip(),
        "name": name,
        "x": x,
        "y": y,
        "is_leave": is_leave,
    }


def normalize_students(raw_students):
    if not isinstance(raw_students, list):
        return []

    students = []
    for raw_student in raw_students:
        student = normalize_student(raw_student)
        if student:
            students.append(student)
    return students


def parse_students_json(raw_value):
    try:
        return json.loads(raw_value or "[]")
    except (TypeError, json.JSONDecodeError):
        return []


def count_grades(students, ths):
    counts = {"A++": 0, "A+": 0, "A": 0, "B++": 0}
    for student in normalize_students(students):
        if student["is_leave"]:
            continue
        grade = grade_for_score(total_score(student["x"], student["y"]), ths)
        if grade in counts:
            counts[grade] += 1
    return counts


def split_exam_name(exam_name):
    parts = exam_name.strip().split()
    if len(parts) >= 3:
        return [parts[0], " ".join(parts[1:-1]), parts[-1]]
    return ["", exam_name.strip() or "成績報表", ""]


def safe_download_name(name, default="成績報表"):
    cleaned = re.sub(r'[\\/:*?"<>|\r\n]+', "_", name).strip(" ._")
    return cleaned or default


def read_students_initial(file_stream):
    wb = None
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        ws = wb.active
        students = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 10:
                continue

            name = str(row[5]).strip() if row[5] is not None else ""
            student_id = str(row[4]).strip() if row[4] is not None else ""
            if name in {"", "預設標準答案", "None"}:
                continue

            students.append(
                {
                    "id": student_id,
                    "name": name,
                    "x": clamp(to_float(row[9]), 0.0, X_MAX),
                    "y": 0.0,
                    "is_leave": False,
                }
            )
        return students
    except (InvalidFileException, OSError, ValueError) as exc:
        LOGGER.warning("Failed to read uploaded workbook: %s", exc)
        return []
    finally:
        if wb is not None:
            wb.close()


def build_excel(students_data, exam_lines, ths):
    normal_list = []
    leave_list = []

    for student in normalize_students(students_data):
        if student["is_leave"]:
            leave_list.append({"name": student["name"], "is_leave": True})
            continue

        total = round(total_score(student["x"], student["y"]), 2)
        normal_list.append(
            {
                "name": student["name"],
                "x": student["x"],
                "y": student["y"],
                "total": total,
                "is_leave": False,
            }
        )

    sorted_normal = sorted(normal_list, key=lambda student: -student["total"])
    final_students = sorted_normal + leave_list

    n_total = len(final_students)
    n_normal = len(sorted_normal)
    counts = count_grades(sorted_normal, ths)

    rows_per_block = math.ceil((n_total + 2) / 3) if n_total > 0 else 1
    header_row, data_start = 1, 2
    final_row = data_start + rows_per_block - 1

    wb = openpyxl.Workbook()
    ws = wb.active

    apply_report_column_widths(ws)

    for block in range(3):
        base = block * 4 + 1
        for idx, header in enumerate(["姓名", "選擇", "非選", "總分"]):
            sc(ws, header_row, base + idx, header, border=all_thin())

    for idx, student in enumerate(final_students):
        block, row = idx // rows_per_block, data_start + (idx % rows_per_block)
        col = block * 4 + 1

        if student["is_leave"]:
            sc(ws, row, col, student["name"], border=all_thin())
            ws.merge_cells(start_row=row, start_column=col + 1, end_row=row, end_column=col + 3)
            sc(ws, row, col + 1, "假", border=all_thin())
            for offset in range(1, 4):
                ws.cell(row=row, column=col + offset).border = all_thin()
            continue

        grade = grade_for_score(student["total"], ths)
        fill = FILLS.get(grade)
        for offset, value in enumerate([student["name"], student["x"], student["y"], student["total"]]):
            sc(ws, row, col + offset, value, fill=fill, border=all_thin())

    if n_total > 0:
        avg_pos = n_total
        avg_block, avg_start_row = avg_pos // rows_per_block, data_start + (avg_pos % rows_per_block)
        avg_col_base = avg_block * 4 + 1

        if n_normal > 0:
            avg_vals = [
                "平均",
                round(sum(student["x"] for student in sorted_normal) / n_normal, 2),
                round(sum(student["y"] for student in sorted_normal) / n_normal, 2),
                round(sum(student["total"] for student in sorted_normal) / n_normal, 2),
            ]
        else:
            avg_vals = ["平均", 0, 0, 0]

        for offset, value in enumerate(avg_vals):
            curr_col = avg_col_base + offset
            for fill_row in range(avg_start_row, final_row + 1):
                sc(ws, fill_row, curr_col, "", size=16, border=all_thin())
            if final_row > avg_start_row:
                ws.merge_cells(
                    start_row=avg_start_row,
                    start_column=curr_col,
                    end_row=final_row,
                    end_column=curr_col,
                )
            sc(ws, avg_start_row, curr_col, value, bold=True, size=16, border=all_thin())

    title_r1, title_r2, title_c1, title_c2 = 2, 7, 14, 16
    ws.merge_cells(start_row=title_r1, start_column=title_c1, end_row=title_r2, end_column=title_c2)
    title_cell = ws.cell(row=title_r1, column=title_c1)
    title_cell.value = "\n".join(exam_lines)
    title_cell.font = Font(name=FONT_NAME, bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(title_r1, title_r2 + 1):
        for col in range(title_c1, title_c2 + 1):
            ws.cell(row=row, column=col).border = outer_med(row, col, title_r1, title_c1, title_r2, title_c2)

    grade_r1 = title_r2 + 4
    for idx, (grade, count) in enumerate(counts.items()):
        row = grade_r1 + idx
        fill = FILLS.get(grade)
        for col, value in [(14, grade), (15, count)]:
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(name=FONT_NAME, bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = outer_med(row, col, grade_r1, 14, grade_r1 + 3, 15)
            if fill:
                cell.fill = fill

    apply_report_column_widths(ws)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf


# UI template
HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>TIME SAVER</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <style>
        body {
            background-color: #020617;
            background-image: radial-gradient(at 0% 0%, #1e1b4b 0, transparent 50%), radial-gradient(at 100% 0%, #312e81 0, transparent 50%), radial-gradient(at 50% 100%, #0f172a 0, transparent 50%);
            background-attachment: fixed; min-height: 100vh; color: #f8fafc; font-family: system-ui, -apple-system, sans-serif;
        }
        .glass-card { background: rgba(15, 23, 42, 0.6); backdrop-filter: blur(12px); border: 1px solid rgba(255, 255, 255, 0.1); }
        .btn-main { background: linear-gradient(135deg, #6366f1 0%, #a855f7 100%); transition: all 0.3s; }
        .btn-main:hover:not(:disabled) { transform: translateY(-2px); box-shadow: 0 0 20px rgba(168, 85, 247, 0.4); }
        .custom-scrollbar::-webkit-scrollbar { width: 6px; }
        .custom-scrollbar::-webkit-scrollbar-thumb { background: #334155; border-radius: 10px; }
    </style>
</head>
<body class="p-4 md:p-12 flex justify-center">
    <div class="max-w-4xl w-full space-y-8">
        <header class="text-center"><h1 class="text-4xl font-extrabold tracking-tighter bg-gradient-to-r from-indigo-400 via-purple-400 to-pink-400 bg-clip-text text-transparent">TIME SAVER</h1></header>

        <div class="glass-card rounded-2xl p-8 space-y-6 shadow-2xl">
            <form id="main-form" action="/generate" method="post" class="space-y-6">
                <input type="hidden" id="students-json" name="students_json" value="">

                <div>
                    <label class="text-[10px] uppercase tracking-[0.2em] text-slate-500 font-bold mb-2 block">考試名稱</label>
                    <input type="text" name="exam_name" placeholder="例如：國三 金安模擬考 第一回" class="w-full bg-slate-950/50 border border-slate-800 rounded-lg px-4 py-3 focus:ring-2 focus:ring-indigo-500 outline-none">
                </div>

                <div class="grid grid-cols-2 md:grid-cols-4 gap-4 text-center">
                    <div class="space-y-1"><label class="text-[10px] text-slate-500 font-bold">A++ 門檻</label><input type="number" step="0.1" id="th_app" name="th_app" value="93.2" class="th-input w-full bg-slate-950/50 border border-slate-800 rounded-lg p-2 text-center text-indigo-400"></div>
                    <div class="space-y-1"><label class="text-[10px] text-slate-500 font-bold">A+ 門檻</label><input type="number" step="0.1" id="th_ap" name="th_ap" value="85.7" class="th-input w-full bg-slate-950/50 border border-slate-800 rounded-lg p-2 text-center text-emerald-400"></div>
                    <div class="space-y-1"><label class="text-[10px] text-slate-500 font-bold">A 門檻</label><input type="number" step="0.1" id="th_a" name="th_a" value="76.2" class="th-input w-full bg-slate-950/50 border border-slate-800 rounded-lg p-2 text-center text-amber-400"></div>
                    <div class="space-y-1"><label class="text-[10px] text-slate-500 font-bold">B++ 門檻</label><input type="number" step="0.1" id="th_bpp" name="th_bpp" value="67.1" class="th-input w-full bg-slate-950/50 border border-slate-800 rounded-lg p-2 text-center text-pink-400"></div>
                </div>

                <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
                    <div id="dropzone" class="border-2 border-dashed border-slate-800 rounded-xl p-8 text-center hover:border-indigo-500 cursor-pointer">
                        <input type="file" id="file-input" accept=".xlsx,.xlsm" class="hidden">
                        <div class="text-2xl mb-1">檔案</div><p class="text-xs text-slate-500">讀卡機 XLSX 或 XLSM 檔案</p>
                    </div>
                    <div class="border border-slate-800 rounded-xl p-4 bg-slate-900/30 space-y-3">
                        <p class="text-[10px] text-indigo-400 font-bold uppercase">手動新增</p>
                        <div class="flex flex-wrap gap-2 items-center">
                            <input type="text" id="manual-name" placeholder="姓名" class="flex-1 min-w-[80px] bg-slate-950 border border-slate-700 rounded px-2 py-1 text-sm outline-none">
                            <input type="number" id="manual-x" min="0" max="25" step="0.5" placeholder="選擇" class="w-16 bg-slate-950 border border-slate-700 rounded px-2 py-1 text-sm text-center">
                            <label class="flex items-center gap-1 text-xs text-slate-400 cursor-pointer">
                                <input type="checkbox" id="manual-leave" class="accent-indigo-500"> 請假
                            </label>
                            <button type="button" id="manual-add-btn" class="bg-indigo-600 hover:bg-indigo-500 px-3 py-1 rounded text-xs font-bold transition-colors">新增</button>
                        </div>
                    </div>
                </div>

                <div id="success-bar" class="hidden bg-indigo-500/10 border border-indigo-500/20 py-3 px-4 rounded-lg flex items-center justify-between">
                    <span class="text-xs text-indigo-300 font-medium">載入成功：<span id="st-count">0</span> 位學生</span>
                    <span class="text-[10px] px-2 py-0.5 bg-indigo-500/20 rounded-full text-indigo-400 font-bold uppercase">Ready</span>
                </div>

                <div id="students-scores-container" class="hidden space-y-4">
                    <div class="flex justify-between items-center border-b border-slate-800 pb-2">
                        <h3 class="text-sm font-bold text-indigo-400">登錄非選分數 (滿分 6)</h3>
                        <span class="text-[10px] text-slate-500 italic">總分 = (X/25)*85 + (Y/6)*15</span>
                    </div>
                    <div id="students-list" class="max-h-64 overflow-y-auto pr-1 space-y-2 custom-scrollbar"></div>
                </div>

                <div class="space-y-2">
                    <label class="text-[10px] uppercase tracking-[0.2em] text-indigo-400 font-bold block">名單順序 (每人一行)</label>
                    <textarea id="ordered_names" name="ordered_names" rows="5" placeholder="請直接貼入補習班系統的名單順序..." class="w-full bg-slate-950/50 border border-slate-800 rounded-lg px-4 py-3 text-sm focus:ring-2 focus:ring-indigo-500 outline-none custom-scrollbar"></textarea>
                </div>

                <div class="flex flex-col md:flex-row gap-4 pt-4">
                    <button type="submit" id="submit-btn" class="btn-main flex-[2] py-4 rounded-xl font-bold text-white shadow-xl opacity-50 cursor-not-allowed" disabled>下載正式成績報表 (.xlsx)</button>
                    <button type="button" id="copy-btn" class="flex-1 bg-slate-800 border border-slate-700 hover:bg-slate-700 text-slate-200 py-4 rounded-xl font-bold opacity-50 cursor-not-allowed" disabled>生成 APP 貼上表</button>
                </div>
            </form>
        </div>

        <div class="grid grid-cols-2 md:grid-cols-4 gap-4">
            <div class="glass-card rounded-2xl p-6 text-center"><p class="text-[10px] text-indigo-400 font-bold">A++</p><h2 id="sum-app" class="text-3xl font-black">--</h2></div>
            <div class="glass-card rounded-2xl p-6 text-center"><p class="text-[10px] text-emerald-400 font-bold">A+</p><h2 id="sum-ap" class="text-3xl font-black">--</h2></div>
            <div class="glass-card rounded-2xl p-6 text-center"><p class="text-[10px] text-amber-400 font-bold">A</p><h2 id="sum-a" class="text-3xl font-black">--</h2></div>
            <div class="glass-card rounded-2xl p-6 text-center"><p class="text-[10px] text-pink-400 font-bold">B++</p><h2 id="sum-bpp" class="text-3xl font-black">--</h2></div>
        </div>
    </div>

    <script>
        const fileInput = document.getElementById('file-input');
        const dropzone = document.getElementById('dropzone');
        const studentsJsonInput = document.getElementById('students-json');
        let studentsData = [];

        function clampScore(value, min, max) {
            const numberValue = Number.parseFloat(value);
            if (!Number.isFinite(numberValue)) return min;
            return Math.min(max, Math.max(min, numberValue));
        }

        function createEl(tagName, className, text) {
            const element = document.createElement(tagName);
            if (className) element.className = className;
            if (text !== undefined) element.textContent = text;
            return element;
        }

        function syncStudentsJson() {
            studentsJsonInput.value = JSON.stringify(studentsData);
        }

        dropzone.addEventListener('click', () => fileInput.click());

        fileInput.addEventListener('change', async () => {
            const file = fileInput.files[0];
            if (!file) return;

            const formData = new FormData();
            formData.append('file', file);

            try {
                const res = await fetch('/upload_read', { method: 'POST', body: formData });
                if (!res.ok) throw new Error(`Upload failed: ${res.status}`);
                const data = await res.json();

                if (Array.isArray(data.students) && data.students.length > 0) {
                    data.students.forEach((student) => {
                        const name = String(student.name || '').trim();
                        if (name && !studentsData.some((current) => current.name === name)) {
                            studentsData.push({
                                name,
                                x: clampScore(student.x, 0, 25),
                                y: 0,
                                is_leave: false
                            });
                        }
                    });
                    refreshUI();
                    return;
                }

                alert('檔案讀取失敗或格式不正確（請確保學生姓名在第 F 欄，選擇分數在第 J 欄）');
            } catch (error) {
                console.error(error);
                alert('檔案上傳或讀取時發生錯誤。');
            }
        });

        function addManualStudent() {
            const nameInp = document.getElementById('manual-name');
            const xInp = document.getElementById('manual-x');
            const leaveInp = document.getElementById('manual-leave');

            const name = nameInp.value.trim();
            const isLeave = leaveInp.checked;
            const x = isLeave ? 0 : clampScore(xInp.value, 0, 25);

            if (!name) return;
            if (studentsData.some((student) => student.name === name)) {
                alert('名單中已經有這位學生。');
                return;
            }

            studentsData.push({ name, x, y: 0, is_leave: isLeave });

            nameInp.value = '';
            xInp.value = '';
            leaveInp.checked = false;
            refreshUI();
        }

        function removeStudent(index) {
            studentsData.splice(index, 1);
            refreshUI();
        }

        function refreshUI() {
            const hasData = studentsData.length > 0;
            const submitBtn = document.getElementById('submit-btn');
            const copyBtn = document.getElementById('copy-btn');
            syncStudentsJson();

            if (hasData) {
                document.getElementById('st-count').innerText = studentsData.length;
                document.getElementById('success-bar').classList.remove('hidden');
                [submitBtn, copyBtn].forEach((btn) => {
                    btn.disabled = false;
                    btn.classList.remove('opacity-50', 'cursor-not-allowed');
                });
                renderStudentsInputs();
                updateDashboard();
            } else {
                document.getElementById('success-bar').classList.add('hidden');
                document.getElementById('students-scores-container').classList.add('hidden');
                [submitBtn, copyBtn].forEach((btn) => {
                    btn.disabled = true;
                    btn.classList.add('opacity-50', 'cursor-not-allowed');
                });
            }
        }

        function renderStudentsInputs() {
            const listDiv = document.getElementById('students-list');
            listDiv.textContent = '';

            studentsData.forEach((student, idx) => {
                const row = createEl('div', 'flex items-center justify-between bg-slate-950/40 border border-slate-800/80 px-4 py-2 rounded-lg gap-4');
                const details = createEl('div', 'flex-1 flex items-center justify-between');
                const name = createEl('span', 'text-sm font-semibold text-slate-200', student.name);
                details.appendChild(name);

                if (student.is_leave) {
                    details.appendChild(createEl('span', 'text-xs font-bold text-slate-500 bg-slate-800 px-2 py-0.5 rounded', '請假'));
                } else {
                    const score = createEl('span', 'text-[10px] text-slate-500');
                    score.append('選擇: ');
                    score.appendChild(createEl('b', 'text-indigo-300', String(student.x)));
                    details.appendChild(score);
                }

                const actions = createEl('div', 'flex items-center gap-3');
                if (!student.is_leave) {
                    const yInput = createEl('input', 'student-y-input w-12 bg-slate-950 border border-slate-700 rounded p-1 text-center text-xs font-bold text-emerald-400 outline-none');
                    yInput.type = 'number';
                    yInput.min = '0';
                    yInput.max = '6';
                    yInput.step = '0.5';
                    yInput.value = student.y;
                    yInput.addEventListener('input', () => {
                        student.y = clampScore(yInput.value, 0, 6);
                        syncStudentsJson();
                        updateDashboard();
                    });
                    actions.appendChild(yInput);
                } else {
                    actions.appendChild(createEl('div', 'w-12'));
                }

                const removeBtn = createEl('button', 'text-slate-600 hover:text-red-500 text-lg transition-colors', '×');
                removeBtn.type = 'button';
                removeBtn.addEventListener('click', () => removeStudent(idx));
                actions.appendChild(removeBtn);

                row.appendChild(details);
                row.appendChild(actions);
                listDiv.appendChild(row);
            });

            document.getElementById('students-scores-container').classList.remove('hidden');
        }

        async function updateDashboard() {
            syncStudentsJson();
            if (studentsData.length === 0) return;

            const payload = {
                students: studentsData,
                th_app: Number.parseFloat(document.getElementById('th_app').value) || 0,
                th_ap: Number.parseFloat(document.getElementById('th_ap').value) || 0,
                th_a: Number.parseFloat(document.getElementById('th_a').value) || 0,
                th_bpp: Number.parseFloat(document.getElementById('th_bpp').value) || 0
            };

            try {
                const res = await fetch('/analyze_full', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(payload)
                });
                if (!res.ok) throw new Error(`Analyze failed: ${res.status}`);
                const data = await res.json();
                document.getElementById('sum-app').innerText = data.counts['A++'];
                document.getElementById('sum-ap').innerText = data.counts['A+'];
                document.getElementById('sum-a').innerText = data.counts['A'];
                document.getElementById('sum-bpp').innerText = data.counts['B++'];
            } catch (error) {
                console.error(error);
            }
        }

        function downloadCopyList() {
            const names = document.getElementById('ordered_names').value.trim();
            if (!names) {
                alert('請先貼入 APP 名單順序！');
                return;
            }

            syncStudentsJson();
            const form = document.getElementById('main-form');
            const originalAction = form.action;
            form.action = '/generate_copy_list';
            form.submit();
            setTimeout(() => { form.action = originalAction; }, 500);
        }

        document.getElementById('manual-add-btn').addEventListener('click', addManualStudent);
        document.getElementById('copy-btn').addEventListener('click', downloadCopyList);
        document.getElementById('main-form').addEventListener('submit', syncStudentsJson);
        document.querySelectorAll('.th-input').forEach((input) => input.addEventListener('input', updateDashboard));
    </script>
</body>
</html>'''


@app.route("/")
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route("/upload_read", methods=["POST"])
def upload_read():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"students": []}), 400
    return jsonify({"students": read_students_initial(io.BytesIO(file.read()))})


@app.route("/analyze_full", methods=["POST"])
def analyze_full():
    data = request.get_json(silent=True) or {}
    students = normalize_students(data.get("students", []))
    ths = threshold_values(data)
    return jsonify({"counts": count_grades(students, ths)})


@app.route("/generate", methods=["POST"])
def generate():
    exam_name = request.form.get("exam_name", "").strip() or "成績報表"
    ths = threshold_values(request.form)
    students = normalize_students(parse_students_json(request.form.get("students_json", "[]")))
    if not students:
        return "沒有學生資料", 400

    filename = f"{safe_download_name(exam_name)}.xlsx"
    return send_file(
        build_excel(students, split_exam_name(exam_name), ths),
        as_attachment=True,
        download_name=filename,
        mimetype=XLSX_MIMETYPE,
    )


@app.route("/generate_copy_list", methods=["POST"])
def generate_copy_list():
    students = normalize_students(parse_students_json(request.form.get("students_json", "[]")))
    ordered_names_raw = request.form.get("ordered_names", "")
    ordered_names = [name.strip() for name in ordered_names_raw.splitlines() if name.strip()]

    student_map = {}
    for student in students:
        student_map.setdefault(student["name"], student)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "補習班貼上專用"
    ws.cell(row=1, column=1, value="APP名單順序")
    ws.cell(row=1, column=2, value="總分 (表現)")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14

    for row, target_name in enumerate(ordered_names, start=2):
        ws.cell(row=row, column=1, value=target_name)
        student = student_map.get(target_name)
        if not student:
            ws.cell(row=row, column=2, value="")
        elif student["is_leave"]:
            ws.cell(row=row, column=2, value="假")
        else:
            total = total_score(student["x"], student["y"])
            ws.cell(row=row, column=2, value=round(total, 2) if total > 0 else "")

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="App_Score_Import.xlsx",
        mimetype=XLSX_MIMETYPE,
    )


if __name__ == "__main__":
    app.run(debug=True)
